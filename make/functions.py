# -*- coding: utf-8 -*-
import openpyxl
import json
import numpy as np
import neal
import ast
import itertools

def CellList(excel_file, sheet_num=0):
    """
    入力：エクセルファイルのpath
    出力：[[ID, NAME, CLASS, TEACHER, PLACE], 
          [ID, NAME, CLASS, TEACHER, PLACE],
             ... 
          [ID, NAME, CLASS, TEACHER, PLACE]]
    """
    def get_sheet(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        sheet_name = wb.sheetnames[sheet_num]
        return wb[sheet_name]
    sheet = get_sheet(excel_file)
    cell_list = list(sheet.values)
    return cell_list[1:]

def ClassDict(cell_list):
    """
    入力：cell_list
    出力：class_dict = {ID : {'NAME':'J1', 'CLASS':['1A'], 'TEACHER':['a'], 'PLACE':['1A']},
                       ID : {'NAME':'J2', 'CLASS':['1A'], 'TEACHER':['b'], 'PLACE':['1A']},
                            ...
                       ID : {'NAME':'PE3', 'CLASS':['1A','1B'], 'TEACHER':['g'], 'PLACE':['体育館']}}
    """
    class_dict = {}
    total = len(cell_list)
    for i in range(total):
        class_dict[i] = {}
        class_dict[i]['NAME'] = cell_list[i][1]
        if cell_list[i][2]:
            class_dict[i]['CLASS'] = list(cell_list[i][2].split(','))
        else:
            class_dict[i]['CLASS'] = []
        class_dict[i]['TEACHER'] = list(cell_list[i][3].split(','))
        if cell_list[i][4]:
            class_dict[i]['PLACE'] = list(cell_list[i][4].split(','))
        else:
            class_dict[i]['PLACE'] = []
    return class_dict

def Adjacent(class_dict):
    """
    入力：class_dict
    出力：adjacent (隣接、重なってはいけない授業)
    """
    total = len(class_dict) # cell_listには全ての授業の情報が入っているはず
    adjacent = {}
    for i in range(total):
        adjacent[i] = []
        for j in range(total):
            if i != j:
                c1 = class_dict[i]
                c2 = class_dict[j]
                # CLASS or TEACHER or PLACEが同じ(もしくは被ってるものがある)
                if set(c1['CLASS'])&set(c2['CLASS']) or set(c1['TEACHER'])&set(c2['TEACHER']) or set(c1['PLACE'])&set(c2['PLACE']):
                    adjacent[i].append(j)
    return adjacent

def Hamiltonian(class_dict,
                weekly, total,
                adjacent,  # 第1項
                convenience,  # 第3項
                renzoku_koma, renzoku_ID, # 第4項
                table, one_per_day,  # 第5項
                joint,  # 第6項
                gen_list, one_per_gen,   # 第7項
                renzoku_2koma, not_renzoku_ID,  # 第8項
                w1=1.0, w2=1.0, w3=1.0, w4=1.0, w5=1.0, w6=1.0, w7=1.0, w8=1.0  # 重み
                ):
    """
    ハミルトニアンを計算してQUBOとして返す
    q: 1週間のコマの数、つまりweekly
    alpha: 制約の強さ
    """
    constant = 0
    A = np.zeros((total*weekly, total*weekly))
    # 第1項：問題を表す
    for i in range(total):
        for j in adjacent[i]:
            if sorted([i, j]) in joint:
                continue
            for a in range(weekly):
                k1 = weekly * i + a
                k2 = weekly * j + a
                A[k1, k2] += w1
    # 第2項：制約条件（１つの科目に１つのコマしか割り当てない）
    plus2 = 0
    for i in range(total):
        if len(class_dict[i]['CLASS']) > 1:
            plus2 = 0
        else:
            plus2 = 0
        for a in range(weekly):
            k1 = weekly * i + a
            for b in range(weekly):
                k2 = weekly * i + b
                if k1 == k2: # 1次の項は対角成分に入れる
                    A[k1, k2] -= (w2 + plus2)
                else:
                    A[k1, k2] += (w2 + plus2)
    # 第3項：制約条件（教員の都合を反映）
    for con in convenience:
        # convenience：[[ID, c], [ID, c], ...]
        i, a = con[0], con[1]
        k = weekly * i + a
        A[k, k] += w3
    # 第4項：制約条件（2時間続きにしたい授業）
    for IDs in renzoku_ID:
        constant += w4
        for renzoku in renzoku_koma:
            k1 = weekly * IDs[0] + renzoku[0]
            k2 = weekly * IDs[1] + renzoku[1]
            A[k1, k2] -= w4
    # 第5項：制約条件（各科目は1日1コマ）
    for IDs in one_per_day:
        for komas in table:
            for i in IDs:
                for j in IDs:
                    for a in komas:
                        for b in komas:
                            k1 = weekly * i + a
                            k2 = weekly * j + b
                            if k1 != k2:
                                A[k1, k2] += w5
    # 第6項：制約条件（2クラスでの合同授業）
    """
    TODO: 第1項との絡みをどうするか、(同じ先生、同じ場所の科目はそもそも繋ぐようにしている)
    合同にしたい授業はグラフ上でそもそも繋がないか。
    また、3クラス以上で可能なのか
    """
    for IDs in joint:
        constant += w6
        for a in range(weekly):
            k1 = weekly * IDs[0] + a
            k2 = weekly * IDs[1] + a
            A[k1, k2] -= w6
    # 第7項：制約条件（各科目が同じ時間帯に固まらないようにする）
    for IDs in one_per_gen:
        for komas in gen_list:
            for i in IDs:
                for j in IDs:
                    for a in komas:
                        for b in komas:
                            k1 = weekly * i + a
                            k2 = weekly * j + b
                            if k1 != k2:
                                A[k1, k2] += w7
    # 第8項：制約条件（1教員が3コマ連続にならないようにする）
    for ID_list in not_renzoku_ID:
        ID2 = list(itertools.combinations(ID_list, 2))
        for IDs in ID2:
            for koma2 in renzoku_2koma:
                k1 = weekly * IDs[0] + koma2[0]
                k2 = weekly * IDs[1] + koma2[1]
                A[k1, k2] += w8
                
    # dimodソルバー用のQUBO形式に変換
    Q = {}
    for k1 in range(total*weekly):
        for k2 in range(total*weekly):
            if A[k1, k2] != 0:
                Q[(k1, k2)] = A[k1, k2]
    return Q, constant, A

def search(table, koma):
    for i in range(len(table)):
        j = 0
        for t in table[i]:
            if koma == t:
                return i, j
            j += 1
            
def ClassName(class_dict):
    total = len(class_dict)
    class_name = set()
    for i in range(total):
        for name in class_dict[i]['CLASS']:
            class_name.add(name)
    class_list = list(class_name)
    class_list.sort()
    return class_list

def TeacherName(class_dict):
    total = len(class_dict)
    teacher_name = set()
    for i in range(total):
        for name in class_dict[i]['TEACHER']:
            teacher_name.add(name)
    teacher_list = list(teacher_name)
    teacher_list.sort()
    return teacher_list

def ClassTable(koma_data, class_dict, class_name, table):
    class_table = {}
    for name in class_name:
        # tableの形によるかも
        class_table[name] = [[""] * len(table) for i in range(len(table[0]))]
    total = len(class_dict)
    #for ID in range(total):
    for ID in koma_data:
        x, y = search(table, koma_data[ID])
        for name in class_dict[ID]['CLASS']:
            # ここでx, yをひっくり返している
            class_table[name][y][x] = class_dict[ID]['NAME'] + ',' + ','.join(class_dict[ID]['TEACHER']) + ',' + ','.join(class_dict[ID]['PLACE'])
    return class_table

def TeacherTable(koma_data, class_dict, teacher_name, weekly):
    teacher_table = {}
    for name in teacher_name:
        teacher_table[name] = [[None] * weekly]
    total = len(class_dict)
    #for ID in range(total):
    for ID in koma_data:
        order = koma_data[ID]
        for name in class_dict[ID]['TEACHER']:
            teacher_table[name][0][order] = class_dict[ID]['NAME'] + ',' + ','.join(class_dict[ID]['CLASS'])
    return teacher_table

# --- 時間割を別のExcelファイルに出力 ---
def write_list_2d(sheet, list2d, start_row, start_col):
    """
    2次元配列を'そのまま見たまんま'Excelファイルのセルに書き込む
    入力：sheet, 2次元配列, 左上セルの行, 左上セルの列
    """
    for y, row in enumerate(list2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                      column=start_col + x,
                      value=list2d[y][x])
            
def MakeExcelFile(output_file, class_dict, koma_data, table, weekly):
    # 書き込みに必要なデータ
    class_name = ClassName(class_dict)
    class_table = ClassTable(koma_data, class_dict, class_name, table)
    teacher_name = TeacherName(class_dict)
    teacher_table = TeacherTable(koma_data, class_dict, teacher_name, weekly)

    # 書き込み用
    wb = openpyxl.Workbook()

    # 曜日、何限目などの体裁を整える
    weekday = ['月', '火', '水', '木', '金', '土']
    gen = [[1], [2], [3], [4], [5], [6], [7], [8], [9]]

    # sheet1
    sheet1 = wb.create_sheet('クラス別の時間割')
    for i in range(len(class_name)):
        # クラス名
        write_list_2d(sheet1, [[class_name[i]]], 1+i*(len(table[0])+2), 1)
        # 曜日
        write_list_2d(sheet1, [weekday[:len(table)]], 1+i*(len(table[0])+2), 2)
        # 何限
        write_list_2d(sheet1, gen[:len(table[0])], 2+i*(len(table[0])+2), 1)
        # 時間割
        write_list_2d(sheet1, class_table[class_name[i]], 2+i*(len(table[0])+2),2)
        i += 1
    # sheet2
    sheet2 = wb.create_sheet('先生別の時間割')
    # 曜日・何限
    allkoma = [[]]
    for i in range(len(table)):
        for j in range(len(table[i])):
            string = weekday[i] + str(gen[j][0])
            allkoma[0].append(string)
    write_list_2d(sheet2, allkoma, 1, 2)
    # 先生の時間割
    for i in range(len(teacher_name)):
        write_list_2d(sheet2, [['先生']], 1, 1)
        # 先生の名前
        write_list_2d(sheet2, [[teacher_name[i]]], 2+i, 1)
        # 時間割
        write_list_2d(sheet2, teacher_table[teacher_name[i]], 2+i, 2)

    wb.save(output_file)

def KomaDataList(model, class_dict):
    adjacent = Adjacent(class_dict)
    total = len(class_dict)
    table = json.loads(model.table)
    convenience = Convenience(model, class_dict)
    renzoku_koma = Renzoku2Koma(model)
    renzoku_ID = []
    one_per_day = OnePerGen(model, class_dict)
    joint = Joint()
    gen_list = GenList(model)
    # TODO: OnePerDay()
    one_per_gen = one_per_day
    renzoku_2koma = Renzoku2Koma(model)
    renzoku_3koma = Renzoku3Koma(model)
    not_renzoku_ID = NotRenzokuID(model, class_dict)
    perfect = PerfectScore()
    w1, w2, w3, w4, w5, w6, w7, w8 = 2, 4, 2, 4, 1, 6, 1, 5
    Q, constant, A = Hamiltonian(
        class_dict,
        model.weekly, total, 
        adjacent, 
        convenience, 
        renzoku_koma, renzoku_ID,
        json.loads(model.table), one_per_day,
        joint,
        gen_list, one_per_gen,
        renzoku_2koma, not_renzoku_ID,
        w1=w1, w2=w2, w3=w3, w4=w4, w5=w5, w6=w6, w7=w7, w8=w8)
    solver = neal.SimulatedAnnealingSampler()
    response = solver.sample_qubo(Q, num_sweeps=model.steps, num_reads=model.reads)
    koma_data_list = []
    for sample0, energy0 in response.data(fields=['sample', 'energy']):
        koma_data = {}
        for key, val in sample0.items():
            if val == 1:
                i = key // model.weekly
                a = key % model.weekly
                koma_data[i] = a
        # 基本制約を満たしていれば次のステップ(得点の計算)へ
        if is_satisfied_2(koma_data, total) and is_satisfied_1(adjacent, joint, koma_data):
            # 制約を満たさなかった群
            broken3 = is_satisfied_3(convenience, koma_data)
            broken4 = is_satisfied_4(renzoku_ID, renzoku_koma, koma_data)
            broken5 = is_satisfied_5(one_per_day, table, koma_data)
            broken6 = is_satisfied_6(joint, koma_data)
            broken7 = is_satisfied_7(one_per_gen, gen_list, koma_data)
            broken8 = is_satisfied_8(table, not_renzoku_ID, renzoku_3koma, koma_data)
            # 条件の重み
            penalty3 = 5
            penalty4 = 5
            penalty5 = 5
            penalty6 = 5
            penalty7 = 5
            penalty8 = 5
            # 得点の計算
            score_for_students = ScoreForStudents(perfect, broken5, broken7, penalty5, penalty7)
            score_for_teachers = ScoreForTeachers(perfect, broken3, broken8, penalty3, penalty8)
            score_strict = ScoreStrict(perfect, broken4, broken6, penalty4, penalty6)
            sum_of_scores = sum([score_for_students, score_for_teachers, score_strict])
            koma_data_list.append({
                'koma_data': koma_data,
                'sum': sum_of_scores,
                'students': score_for_students,
                'teachers': score_for_teachers,
                'strict': score_strict,
            })
    # koma_dataを得点の高い順に並べかえる
    koma_data_list.sort(key=lambda x: -x['sum'])
    model.koma_data_list = json.dumps(koma_data_list)
    model.save()
    return koma_data_list

def ClassTableList(model):
    cell_list = json.loads(model.cell_list)
    class_dict = ClassDict(cell_list)
    koma_data_list = KomaDataList(model, class_dict)
    class_table_list = []
    for info in koma_data_list:
        koma_data = info['koma_data']
        class_table = ClassTable(koma_data, class_dict, json.loads(model.class_list), json.loads(model.table))
        class_table_list.append(class_table)
    return class_table_list

def Convenience(model, class_dict):
    convenience = []
    con = ast.literal_eval(model.convenience)
    for ID, info in class_dict.items():
        for teacher in con:
            if teacher in info['TEACHER']:
                for koma in con[teacher]:
                    convenience.append([ID, koma])
    return convenience

def OnePerDay():
    return []

def Joint():
    return []

def GenList(model):
    gen_list = [[], []]
    table = json.loads(model.table)
    lunch_after = model.lunch_after
    min_class = 100
    for t in table:
        min_class = min(min_class, len(t))
    for t in table:
        gen_list[0].append(t[0])
        if lunch_after > 0 and lunch_after < min_class:
            gen_list[1].append(t[lunch_after])
    return gen_list

def OnePerGen(model, class_dict):
    classes = {}
    for ID, infomation in class_dict.items():
        info = infomation["NAME"] + str(infomation["CLASS"])
        if info in classes:
            classes[info].append(ID)
        else:
            classes[info] = [ID]
    one_per_gen = []
    for c, class_list in classes.items():
        one_per_gen.append(class_list)
    return one_per_gen

def Renzoku2Koma(model):
    table = json.loads(model.table)
    lunch_after = model.lunch_after
    renzoku_2koma = []
    for day in table:
        for gen in range(len(day) - 1):
            if gen != lunch_after - 1:
                renzoku_2koma.append([day[gen], day[gen + 1]])
    return renzoku_2koma

def Renzoku3Koma(model):
    table = json.loads(model.table)
    lunch_after = model.lunch_after
    renzoku_3koma = []
    for day in table:
        for gen in range(len(day) - 2):
            if gen != lunch_after - 1 and gen != lunch_after - 2:
                renzoku_3koma.append([day[gen], day[gen + 1], day[gen + 2]])
    return renzoku_3koma

def NotRenzokuID(model, class_dict):
    classes = {}
    for ID, infomation in class_dict.items():
        teachers = infomation["TEACHER"]
        for teacher in teachers:
            if teacher in classes:
                classes[teacher].append(ID)
            else:
                classes[teacher] = [ID]
    not_renzoku_ID = []
    for c, class_list in classes.items():
        not_renzoku_ID.append(class_list)
    return not_renzoku_ID

# 制約破り判定
# 第1項：グラフで繋がれた授業が被っていない
def is_satisfied_1(adjacent, joint, koma_data):
    broken = []
    satisfied = True
    for i in adjacent:
        for j in adjacent[i]:
            if koma_data[i] == koma_data[j] and tuple(sorted([i, j])) not in joint:
                satisfied = False
                broken.append((i, j))
    if broken:
        # TODO: print()消す
        print("制約1破り：重なってはいけないのに重なっている授業IDの組")
        print(broken)
    return satisfied

# 第2項：１つの科目に１つのコマしか割り当てない
def is_satisfied_2(koma_data, total):
    if len(koma_data) == total:
        return True
    else:
        # TODO: print()消す
        print("制約2破り：コマが割り当てられていない授業IDの個数")
        print(str(total - len(koma_data)) + "個")
        return False

# 第3項：教員の都合を反映
# 以下第8項まで、jupyterのときと違いsatisfiedではなくbrokenを返すことにする
def is_satisfied_3(convenience, koma_data):
    broken = []
    for con in convenience:
        ID, koma = con[0], con[1]
        if koma_data[ID] == koma:
            broken.append("ID: {}, コマ: {}".format(ID, koma))
    if broken:
        print("制約3破り：都合が反映されていない授業IDとそのコマ")
        print(broken)
    return broken

# 第4項：2時間続きにしたい授業
def is_satisfied_4(renzoku_ID, renzoku_koma, koma_data):
    broken = []
    for IDs in renzoku_ID:
        #if abs(koma_data[IDs[0]] - koma_data[IDs[1]]) != 1:
        if sorted([koma_data[IDs[0]], koma_data[IDs[1]]]) not in renzoku_koma:
            broken.append((IDs[0], IDs[1]))
    if broken:
        print("制約4破り：2時間続きになっていない授業IDの組")
        print(broken)
    return broken

# 第5項：各科目は1日1コマ
def is_satisfied_5(one_per_day, table, koma_data):
    broken = set()
    for IDs in one_per_day:
        for day in table:
            times = 0
            for ID in IDs:
                if koma_data[ID] in day:
                    times += 1
            if times > 1:
                broken.add(tuple(IDs))
    if broken:
        print("制約5破り：1日に1コマ以上入っている授業群")
        print(broken)
    return broken

# 第6項：2クラス合同
def is_satisfied_6(joint, koma_data):
    broken = []
    for j in joint:
        if koma_data[j[0]] != koma_data[j[1]]:
            broken.append(j)
    if broken:
        print("制約6破り：2クラス合同になっていない授業群")
        print(broken)
    return broken

# 第7項：各科目が同じ時間帯に固まらないようにする
def is_satisfied_7(one_per_gen, gen_list, koma_data):
    broken = set()
    for IDs in one_per_gen:
        for day in gen_list:
            times = 0
            for ID in IDs:
                if koma_data[ID] in day:
                    times += 1
            if times > 1:
                broken.add(tuple(IDs))
    if broken:
        print("制約7破り：同じ時間帯に入っている授業群")
        print(broken)
    return broken

# 第8項：1教員が3コマ連続にならないようにする
def is_satisfied_8(table, not_renzoku_ID, renzoku_3koma, koma_data):
    broken = []
    for IDs in not_renzoku_ID:
        ID3s = list(itertools.combinations(IDs, 3))
        for ID in ID3s:
            komas = [koma_data[ID[0]], koma_data[ID[1]], koma_data[ID[2]]]
            komas.sort()
            if komas[1] - komas[0] == 1 and komas[2] - komas[1] == 1:
                for t in table:
                    if komas[0] in t and komas[1] in t and komas[2] in t:
                        broken.append(ID)
                        break
    if broken:
        print("制約8破り：1教員が3コマ連続になっている授業群")
        print(broken)
    return broken

# 制約の個数から基準にする満点を算出
def PerfectScore():
    score = 1000
    return score

# 時間割の点数をそれぞれの評価軸において計算する
def ScoreForStudents(perfect, broken5, broken7, penalty5, penalty7):
    total_penalty = 0
    # 制約5:1日1コマ
    total_penalty += penalty5 * len(broken5)
    # 制約7:同じ時間にかためない
    total_penalty += penalty7 * len(broken7)
    return (perfect - total_penalty) * 100 / perfect

def ScoreForTeachers(perfect, broken3, broken8, penalty3, penalty8):
    total_penalty = 0
    # 制約3:先生の都合
    total_penalty += penalty3 * len(broken3)
    # 制約8:3コマ連続にしない
    total_penalty += penalty8 * len(broken8)
    return (perfect - total_penalty) * 100 / perfect

def ScoreStrict(perfect, broken4, broken6, penalty4, penalty6):
    total_penalty = 0
    # 制約4:2時間連続
    total_penalty += penalty4 * len(broken4)
    # 制約6:2クラス合同
    total_penalty += penalty6 * len(broken6)
    return (perfect - total_penalty) * 100 / perfect